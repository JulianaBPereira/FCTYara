import re
from datetime import datetime
from pathlib import Path

from services.logs_path_service import pasta_logs_padrao


def salvar_log(
    receita: str,
    serial: str,
    resultados: list[dict],
    pasta: Path | None = None,
    formato_txt: bool = True,
    formato_excel: bool = False,
) -> list[Path]:
    """Salva os resultados do teste nos formatos selecionados.

    Nome base: <serial>_DD_MM_YYYY  (ex.: 123456789_04_06_2026)
    Retorna lista com os caminhos dos arquivos gerados.
    """
    if pasta is None:
        pasta = pasta_logs_padrao()

    pasta.mkdir(parents=True, exist_ok=True)

    agora = datetime.now()
    data = agora.strftime("%d_%m_%Y")
    serial_safe = re.sub(r'[\\/:*?"<>|]', "_", serial) if serial else "SEM_SERIAL"

    aprovados = [r for r in resultados if str(r.get("status", "")).upper() == "PASS"]
    resultado_geral = "PASS" if len(aprovados) == len(resultados) and resultados else "FAIL"

    nome_base = f"{serial_safe}_{data}"

    gerados: list[Path] = []

    if formato_txt:
        gerados.append(
            _salvar_txt(pasta, nome_base, receita, serial, agora, resultado_geral, resultados)
        )

    if formato_excel:
        gerados.append(
            _salvar_excel(pasta, nome_base, receita, serial, agora, resultado_geral, resultados)
        )

    return gerados


# ── TXT ───────────────────────────────────────────────────────────────────────

def _proximo_caminho(pasta: Path, nome_base: str, extensao: str) -> Path:
    """Retorna o próximo caminho disponível sem sobrescrever arquivos existentes."""
    caminho = pasta / f"{nome_base}{extensao}"
    contador = 1
    while caminho.exists():
        contador += 1
        caminho = pasta / f"{nome_base}_{contador}{extensao}"
    return caminho


def _salvar_txt(
    pasta: Path,
    nome_base: str,
    receita: str,
    serial: str,
    agora: datetime,
    resultado_geral: str,
    resultados: list[dict],
) -> Path:
    caminho = _proximo_caminho(pasta, nome_base, ".txt")

    with open(caminho, "w", encoding="utf-8") as f:
        f.write(f"Receita  : {receita}\n")
        f.write(f"Serial   : {serial or '—'}\n")
        f.write(f"Data/hora: {agora.strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Resultado: {resultado_geral}\n")
        f.write("-" * 65 + "\n")
        f.write(f"{'TESTE':<28} {'RANGE':<15} {'VALOR':<12} {'STATUS'}\n")
        f.write("-" * 65 + "\n")
        for r in resultados:
            f.write(
                f"{str(r.get('nome', '')):<28} "
                f"{str(r.get('range', '')):<15} "
                f"{str(r.get('valor', '')):<12} "
                f"{r.get('status', '')}\n"
            )

    return caminho


# ── Excel ─────────────────────────────────────────────────────────────────────

def _salvar_excel(
    pasta: Path,
    nome_base: str,
    receita: str,
    serial: str,
    agora: datetime,
    resultado_geral: str,
    resultados: list[dict],
) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    caminho = _proximo_caminho(pasta, nome_base, ".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultado"

    azul = "16345C"
    verde = "16A34A"
    vermelho = "DC2626"
    branco = "FFFFFF"
    cinza_claro = "F5F7FA"

    fonte_cabecalho = Font(name="Segoe UI", bold=True, color=branco, size=11)
    fonte_normal = Font(name="Segoe UI", size=11)
    fonte_pass = Font(name="Segoe UI", bold=True, color=verde, size=11)
    fonte_fail = Font(name="Segoe UI", bold=True, color=vermelho, size=11)
    fill_azul = PatternFill("solid", fgColor=azul)
    fill_cinza = PatternFill("solid", fgColor=cinza_claro)
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esq = Alignment(horizontal="left", vertical="center")

    # ── Cabeçalho de info ─────────────────────────────────────────────────────
    info = [
        ("Receita", receita),
        ("Serial", serial or "—"),
        ("Data/hora", agora.strftime("%d/%m/%Y %H:%M:%S")),
        ("Resultado", resultado_geral),
    ]
    for i, (rotulo, valor) in enumerate(info, start=1):
        cel_r = ws.cell(row=i, column=1, value=rotulo)
        cel_r.font = Font(name="Segoe UI", bold=True, size=11)
        cel_r.alignment = alinhamento_esq

        cel_v = ws.cell(row=i, column=2, value=valor)
        cel_v.font = (
            fonte_pass if (rotulo == "Resultado" and valor == "PASS") else
            fonte_fail if (rotulo == "Resultado" and valor == "FAIL") else
            fonte_normal
        )
        cel_v.alignment = alinhamento_esq

    # ── Linha em branco ───────────────────────────────────────────────────────
    linha_tabela = len(info) + 2

    # ── Cabeçalho da tabela ───────────────────────────────────────────────────
    cabecalhos = ["TESTE", "RANGE", "VALOR", "STATUS"]
    for col, titulo in enumerate(cabecalhos, start=1):
        cel = ws.cell(row=linha_tabela, column=col, value=titulo)
        cel.font = fonte_cabecalho
        cel.fill = fill_azul
        cel.alignment = alinhamento_centro

    # ── Dados ─────────────────────────────────────────────────────────────────
    for idx, r in enumerate(resultados):
        linha = linha_tabela + 1 + idx
        fill_linha = fill_cinza if idx % 2 == 0 else PatternFill()
        status = str(r.get("status", ""))

        valores_col = [
            r.get("nome", ""),
            r.get("range", ""),
            r.get("valor", ""),
            status,
        ]
        for col, val in enumerate(valores_col, start=1):
            cel = ws.cell(row=linha, column=col, value=str(val))
            cel.fill = fill_linha
            cel.alignment = alinhamento_centro if col > 1 else alinhamento_esq
            if col == 4:
                cel.font = (
                    fonte_pass if status == "PASS" else
                    fonte_fail if status == "FAIL" else
                    fonte_normal
                )
            else:
                cel.font = fonte_normal

    # ── Larguras das colunas ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12

    wb.save(caminho)
    return caminho


# ── Listagem de logs ──────────────────────────────────────────────────────────

def listar_logs(pasta: Path | None = None) -> list[dict]:
    """Retorna metadados dos arquivos de log encontrados na pasta.

    Cada item: {"arquivo", "tipo", "resultado", "data_hora", "pasta", "caminho"}
    """
    if pasta is None:
        pasta = pasta_logs_padrao()

    if not pasta.exists():
        return []

    # Padrão: <serial>_DD_MM_YYYY.txt/.xlsx  ou  <serial>_DD_MM_YYYY_N.ext
    _PADRAO = re.compile(r"^(.+)_(\d{2}_\d{2}_\d{4})(?:_\d+)?\.(txt|xlsx)$", re.IGNORECASE)

    registros = []
    arquivos = sorted(
        list(pasta.glob("*.txt")) + list(pasta.glob("*.xlsx")),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    for arq in arquivos:
        m = _PADRAO.match(arq.name)
        if m:
            data_str = m.group(2)
            tipo = m.group(3).upper()
            try:
                dt = datetime.strptime(data_str, "%d_%m_%Y")
                data_hora = dt.strftime("%d/%m/%Y")
            except ValueError:
                data_hora = data_str
        else:
            tipo = arq.suffix.lstrip(".").upper()
            data_hora = _data_modificacao(arq)

        resultado = _resultado_do_conteudo(arq) if tipo == "TXT" else _resultado_do_excel(arq)

        registros.append(
            {
                "arquivo": arq.name,
                "tipo": tipo,
                "resultado": resultado,
                "data_hora": data_hora,
                "pasta": str(pasta),
                "caminho": arq,
            }
        )

    return registros


def _resultado_do_conteudo(caminho: Path) -> str:
    try:
        with open(caminho, encoding="utf-8") as f:
            for linha in f:
                if linha.lower().startswith("resultado"):
                    partes = linha.split(":", 1)
                    if len(partes) == 2:
                        return partes[1].strip().upper()
    except OSError:
        pass
    return "—"


def _resultado_do_excel(caminho: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=6, max_col=2, values_only=True):
            if row[0] and str(row[0]).lower() == "resultado":
                return str(row[1]).upper() if row[1] else "—"
    except Exception:
        pass
    return "—"


def _data_modificacao(caminho: Path) -> str:
    try:
        ts = caminho.stat().st_mtime
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M:%S")
    except OSError:
        return "—"
